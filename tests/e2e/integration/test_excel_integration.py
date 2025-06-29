import pytest
from pathlib import Path
from tempfile import TemporaryDirectory
import openpyxl
from src.infrastructure.excel.excel_client_impl import ExcelClientImpl
from decimal import Decimal


@pytest.mark.asyncio
async def test_excel_client_creates_file(sample_cart_item):
    # Arrange
    with TemporaryDirectory() as temp_dir:
        excel_client = ExcelClientImpl(str(Path(temp_dir) / "test_orders.xlsx"))
        
        # Act
        await excel_client.write_order_to_excel(
            user_id=12345,
            full_name="John Doe",
            address="123 Main St",
            phone="+1234567890",
            payment_method="Card",
            cart_items=[sample_cart_item]
        )
        
        # Assert
        assert excel_client.excel_path.exists()
        
        workbook = openpyxl.load_workbook(excel_client.excel_path)
        worksheet = workbook.active
        
        # Check headers
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = ["Дата", "ID пользователя", "ФИО", "Адрес", "Телефон", "Способ оплаты", "Товары", "Общая сумма"]
        assert headers == expected_headers
        
        # Check data row
        data_row = [cell.value for cell in worksheet[2]]
        assert data_row[1] == 12345  # user_id
        assert data_row[2] == "John Doe"  # full_name
        assert data_row[3] == "123 Main St"  # address
        assert data_row[4] == "+1234567890"  # phone
        assert data_row[5] == "Card"  # payment_method
        assert "iPhone x2" in data_row[6]  # products
        assert data_row[7] == float(Decimal("999.99") * 2)  # total


@pytest.mark.asyncio
async def test_excel_client_appends_to_existing_file(sample_cart_item):
    # Arrange
    with TemporaryDirectory() as temp_dir:
        excel_client = ExcelClientImpl(str(Path(temp_dir) / "test_orders.xlsx"))
        
        # Create first order
        await excel_client.write_order_to_excel(
            user_id=12345,
            full_name="John Doe",
            address="123 Main St",
            phone="+1234567890",
            payment_method="Card",
            cart_items=[sample_cart_item]
        )
        
        # Act - Add second order
        await excel_client.write_order_to_excel(
            user_id=67890,
            full_name="Jane Smith",
            address="456 Oak Ave",
            phone="+0987654321",
            payment_method="Cash",
            cart_items=[sample_cart_item]
        )
        
        # Assert
        workbook = openpyxl.load_workbook(excel_client.excel_path)
        worksheet = workbook.active
        
        # Should have header + 2 data rows
        assert worksheet.max_row == 3
        
        # Check second row data
        second_row = [cell.value for cell in worksheet[3]]
        assert second_row[1] == 67890  # user_id
        assert second_row[2] == "Jane Smith"  # full_name